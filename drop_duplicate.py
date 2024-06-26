import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import colors,Font,Border,Side,PatternFill
frame = pd.DataFrame(pd.read_excel(r'D:\之江\网络架构和设备信息\搬迁项目合同设备信息表19.xlsx','合同设备信息'))
Data = frame.drop_duplicates(subset='资产编号',keep='first',inplace=False)
Data.to_excel(r'D:\之江\网络架构和设备信息\搬迁项目合同设备信息表19.xlsx',sheet_name='合同设备信息',index=False)

wb = load_workbook(r'D:\咪咕文档\虚拟化资源池使用统计\低效利用率虚拟机.xlsx')
ws = wb.active
ft1 = Font(name='微软雅黑',size=10,bold=True,italic=False,color=colors.BLACK)
ft2 = Font(name='微软雅黑',size=10,bold=False,italic=False,color=colors.BLACK)
thin = Side(border_style='thin',color=colors.BLACK)
bord = Border(top=thin,bottom=thin,left=thin,right=thin)
fill = PatternFill(fill_type=None,start_color='FFFFFFFF',end_color='FF000000')

for row in ws['A1:H1']:
    for cell in row:
        cell.font = ft1

for row in ws['A1:H1']:
    for cell in row:
        cell.fill = fill

for row in ws['A2:H1200']:
    for cell in row:
        cell.font = ft2

for row in ws['A1:H1200']:
    for cell in row:
        cell.border = bord

wb.save(r'D:\咪咕文档\虚拟化资源池使用统计\低效利用率虚拟机.xlsx')